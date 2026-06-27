"""
clean_pagos.py
Procesa el extracto de pagos/liquidaciones de TikTok (archivo "income_...")
y genera un Excel "Extracto_YYYYMMDD.xlsx" con tres hojas.

Que hace:
1. Abre una ventana para elegir el archivo (Excel o CSV) descargado de TikTok.
2. Copia la hoja "Order details" tal cual en el nuevo Excel "Extracto_YYYYMMDD".
   (Si se elige un CSV, se considera que TODO el CSV son los "Order details").
3. Crea la hoja "Pago_Pedido" con solo las columnas necesarias, renombradas
   al espanol, y una columna CHECK al principio que comprueba que el pago de
   liquidacion cuadra con el desglose:
       Pago_Liquidacion = Pago_Cliente + Descuento_Plataforma + Gasto_Transporte
                          + Comision_Plataforma + Comision_Afiliado
                          + Reembolso_Cliente + Reembolso_Descuento
       -> "OK" si cuadra,  "REVISAR" si no.
4. Crea la hoja "Resumen_Pagos" con los importes consolidados (sumados) por
   ID_Pago.

Las comprobaciones y los resumenes se hacen con FORMULAS de Excel (CHECK,
SUMIF), de modo que el archivo se recalcula solo si se editan los datos.
"""

import re
import sys
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule


# ----------------------------------------------------------------------
# Configuracion
# ----------------------------------------------------------------------
# Nombre de la hoja de origen (se busca sin distinguir mayusculas ni espacios)
HOJA_ORIGEN = "Order details"

# Columnas a extraer:  nombre en TikTok  ->  nombre interno (espanol)
# El orden de este diccionario es el orden de las columnas en "Pago_Pedido"
# (despues de la columna CHECK).
RENOMBRAR = {
    "Payment ID": "ID_Pago",
    "Statement ID": "ID_Extracto",
    "Currency": "Moneda",
    "Related order ID": "ID_Pedido",
    "Total settlement amount": "Pago_Liquidacion",
    "Customer payment": "Pago_Cliente",
    "Platform discounts": "Descuento_Plataforma",
    "Shipping": "Gasto_Transporte",
    "TikTok Shop commission fee": "Comision_Plataforma",
    "Affiliate Commission": "Comision_Afiliado",
    "Refund subtotal before seller discounts": "Reembolso_Cliente",
    "Refund of seller discounts": "Reembolso_Descuento",
}

# Columnas que son IDENTIFICADORES (texto, no numero)
COLS_TEXTO = {"ID_Pago", "ID_Extracto", "Moneda", "ID_Pedido"}

# Columnas que son IMPORTES (numero, con 2 decimales)
COLS_IMPORTE = [
    "Pago_Liquidacion", "Pago_Cliente", "Descuento_Plataforma",
    "Gasto_Transporte", "Comision_Plataforma", "Comision_Afiliado",
    "Reembolso_Cliente", "Reembolso_Descuento",
]

# Columnas del resumen (consolidado por ID_Pago)
COLS_RESUMEN = ["ID_Pago"] + COLS_IMPORTE

TOLERANCIA = 0.005   # margen para que el CHECK acepte redondeos de centimos
FUENTE = "Arial"


# ----------------------------------------------------------------------
# Utilidades
# ----------------------------------------------------------------------
def normaliza(texto):
    """Quita espacios sobrantes (incluso internos) y pasa a minusculas."""
    return re.sub(r"\s+", " ", str(texto)).strip().lower()


def parse_num(valor):
    """Convierte '12,34', '-4.95', 'EUR 5', '/', '' ... en un numero (float)."""
    if valor is None:
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip()
    if texto in ("", "/", "-"):
        return 0.0
    texto = texto.replace("\u00a0", "").replace("€", "")
    texto = re.sub(r"(?i)eur", "", texto).strip()
    if "," in texto and "." in texto:
        if texto.rfind(",") > texto.rfind("."):
            texto = texto.replace(".", "").replace(",", ".")
        else:
            texto = texto.replace(",", "")
    elif "," in texto:
        texto = texto.replace(",", ".")
    try:
        return float(texto)
    except ValueError:
        return 0.0


def clean_id(valor):
    """Identificador como texto limpio (evita notacion cientifica y '.0')."""
    if valor is None:
        return ""
    texto = str(valor).strip()
    if texto.endswith(".0"):
        texto = texto[:-2]
    return texto


def nombre_libre(base_dir, fecha_str):
    """Devuelve una ruta Extracto_YYYYMMDD.xlsx que no exista (anade _2, _3...)."""
    ruta = base_dir / f"Extracto_{fecha_str}.xlsx"
    if not ruta.exists():
        return ruta
    n = 2
    while True:
        ruta = base_dir / f"Extracto_{fecha_str}_{n}.xlsx"
        if not ruta.exists():
            return ruta
        n += 1


# ----------------------------------------------------------------------
# Lectura del origen -> (cabecera, filas)  todo como texto
# ----------------------------------------------------------------------
def leer_origen(ruta):
    """
    Devuelve (cabecera, filas) de la hoja 'Order details'.
    - Excel: busca la hoja por nombre (sin distinguir mayus/espacios).
    - CSV:   considera todo el archivo como los 'Order details'.
    Todos los valores se devuelven como cadena de texto.
    """
    if ruta.suffix.lower() == ".csv":
        import csv
        with open(ruta, newline="", encoding="utf-8-sig") as f:
            lector = list(csv.reader(f))
        if not lector:
            raise ValueError("El CSV esta vacio.")
        cabecera = [c if c is not None else "" for c in lector[0]]
        filas = [[(v if v is not None else "") for v in fila] for fila in lector[1:]]
        return cabecera, filas

    wb = load_workbook(ruta, data_only=True)
    objetivo = normaliza(HOJA_ORIGEN)
    hoja = next((s for s in wb.sheetnames if normaliza(s) == objetivo), None)
    if hoja is None:
        raise ValueError(
            f"No se encontro la hoja '{HOJA_ORIGEN}'.\n"
            f"Hojas del archivo: {', '.join(wb.sheetnames)}"
        )
    ws = wb[hoja]
    datos = list(ws.iter_rows(values_only=True))
    wb.close()
    if not datos:
        raise ValueError(f"La hoja '{hoja}' esta vacia.")

    cabecera = ["" if c is None else str(c) for c in datos[0]]
    filas = []
    for fila in datos[1:]:
        # saltar filas totalmente vacias
        if all(v is None or str(v).strip() == "" for v in fila):
            continue
        filas.append(["" if v is None else str(v) for v in fila])
    return cabecera, filas


# ----------------------------------------------------------------------
# Hoja 1: copia integra de "Order details"
# ----------------------------------------------------------------------
def escribir_order_details(wb, cabecera, filas):
    ws = wb.active
    ws.title = "Order details"
    ws.append(cabecera)
    for fila in filas:
        ws.append(fila)
    # cabecera en negrita
    for celda in ws[1]:
        celda.font = Font(name=FUENTE, bold=True)
    ws.freeze_panes = "A2"
    return ws


# ----------------------------------------------------------------------
# Hoja 2: Pago_Pedido  (CHECK + 12 columnas renombradas)
# ----------------------------------------------------------------------
def escribir_pago_pedido(wb, cabecera, filas):
    # localizar el indice de cada columna pedida (por nombre normalizado)
    indice = {normaliza(c): i for i, c in enumerate(cabecera)}
    faltan, idx_col = [], {}
    for nombre_tt in RENOMBRAR:
        i = indice.get(normaliza(nombre_tt))
        if i is None:
            faltan.append(nombre_tt)
        else:
            idx_col[nombre_tt] = i
    if faltan:
        raise ValueError(
            "Faltan estas columnas en 'Order details':\n- " + "\n- ".join(faltan)
        )

    ws = wb.create_sheet("Pago_Pedido")

    # --- cabecera:  CHECK + nombres en espanol ---
    cabecera_es = ["CHECK"] + list(RENOMBRAR.values())
    ws.append(cabecera_es)

    # mapa nombre_interno -> letra de columna en esta hoja
    letra = {}
    for j, nombre in enumerate(cabecera_es, start=1):
        letra[nombre] = get_column_letter(j)

    L = lambda n: letra[n]   # nombre interno -> letra de columna

    # --- filas de datos ---
    for fila in filas:
        salida = [None]  # hueco para la columna CHECK (formula, se pone luego)
        for nombre_tt, nombre_es in RENOMBRAR.items():
            bruto = fila[idx_col[nombre_tt]] if idx_col[nombre_tt] < len(fila) else ""
            if nombre_es in COLS_TEXTO:
                salida.append(clean_id(bruto))
            else:
                salida.append(parse_num(bruto))
        ws.append(salida)

    # --- columna CHECK: compara la liquidacion con la suma del desglose ---
    for r in range(2, ws.max_row + 1):
        suma = "+".join(f"{L(n)}{r}" for n in [
            "Pago_Cliente", "Descuento_Plataforma", "Gasto_Transporte",
            "Comision_Plataforma", "Comision_Afiliado",
            "Reembolso_Cliente", "Reembolso_Descuento",
        ])
        ws.cell(row=r, column=1).value = (
            f'=IF(ABS({L("Pago_Liquidacion")}{r}-({suma}))<={TOLERANCIA},'
            f'"OK","REVISAR")'
        )

    _formatear_pago_pedido(ws, letra)
    return ws, letra


def _formatear_pago_pedido(ws, letra):
    ultima = ws.max_row
    # cabecera
    for celda in ws[1]:
        celda.font = Font(name=FUENTE, bold=True)
        celda.alignment = Alignment(horizontal="center")
    # fuente general
    for fila in ws.iter_rows(min_row=2, max_row=ultima):
        for celda in fila:
            celda.font = Font(name=FUENTE)
    # formatos de columna
    for nombre, col in letra.items():
        if nombre in COLS_TEXTO:
            for celda in ws[col][1:]:
                celda.number_format = "@"
        elif nombre in COLS_IMPORTE:
            for celda in ws[col][1:]:
                celda.number_format = "0.00"
    # CHECK centrado + color (verde OK / rojo REVISAR)
    col_check = letra["CHECK"]
    for celda in ws[col_check][1:]:
        celda.alignment = Alignment(horizontal="center")
    rango = f"{col_check}2:{col_check}{ultima}"
    verde = PatternFill("solid", fgColor="C6EFCE")
    rojo = PatternFill("solid", fgColor="FFC7CE")
    ws.conditional_formatting.add(
        rango, CellIsRule(operator="equal", formula=['"OK"'], fill=verde,
                          font=Font(name=FUENTE, color="006100"))
    )
    ws.conditional_formatting.add(
        rango, CellIsRule(operator="equal", formula=['"REVISAR"'], fill=rojo,
                          font=Font(name=FUENTE, bold=True, color="9C0006"))
    )
    # anchos
    anchos = {"CHECK": 10, "ID_Pago": 22, "ID_Extracto": 22, "Moneda": 9,
              "ID_Pedido": 22}
    for nombre, col in letra.items():
        ws.column_dimensions[col].width = anchos.get(nombre, 16)
    ws.freeze_panes = "A2"


# ----------------------------------------------------------------------
# Hoja 3: Resumen_Pagos  (consolidado por ID_Pago, con SUMIF)
# ----------------------------------------------------------------------
def escribir_resumen(wb, letra_pp, filas_pp_ids):
    ws = wb.create_sheet("Resumen_Pagos")
    ws.append(COLS_RESUMEN)

    # IDs unicos conservando el orden de aparicion
    vistos, ids_unicos = set(), []
    for pid in filas_pp_ids:
        if pid not in vistos:
            vistos.add(pid)
            ids_unicos.append(pid)

    col_id_pp = letra_pp["ID_Pago"]          # columna de ID_Pago en Pago_Pedido
    for pid in ids_unicos:
        fila = [pid]
        for nombre in COLS_IMPORTE:
            col_src = letra_pp[nombre]
            fila.append(
                f"=SUMIF('Pago_Pedido'!${col_id_pp}:${col_id_pp},$A{ws.max_row + 1},"
                f"'Pago_Pedido'!{col_src}:{col_src})"
            )
        ws.append(fila)

    # fila TOTAL (suma de todo el resumen, para cuadrar)
    primera, ultima = 2, ws.max_row
    ws.append(["TOTAL"] + [
        f"=SUM({get_column_letter(c)}{primera}:{get_column_letter(c)}{ultima})"
        for c in range(2, len(COLS_RESUMEN) + 1)
    ])
    fila_total = ws.max_row

    _formatear_resumen(ws, fila_total)
    return ws


def _formatear_resumen(ws, fila_total):
    for celda in ws[1]:
        celda.font = Font(name=FUENTE, bold=True)
        celda.alignment = Alignment(horizontal="center")
    for fila in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for celda in fila:
            celda.font = Font(name=FUENTE)
    # ID_Pago como texto, importes con 2 decimales
    for celda in ws["A"][1:]:
        celda.number_format = "@"
    for col in range(2, len(COLS_RESUMEN) + 1):
        for celda in ws[get_column_letter(col)][1:]:
            celda.number_format = "0.00"
    # fila TOTAL en negrita
    for celda in ws[fila_total]:
        celda.font = Font(name=FUENTE, bold=True)
    ws.column_dimensions["A"].width = 22
    for col in range(2, len(COLS_RESUMEN) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.freeze_panes = "A2"


# ----------------------------------------------------------------------
# Proceso completo (separado de la GUI para poder probarlo)
# ----------------------------------------------------------------------
def procesar(ruta):
    cabecera, filas = leer_origen(ruta)

    wb = Workbook()
    escribir_order_details(wb, cabecera, filas)
    ws_pp, letra_pp = escribir_pago_pedido(wb, cabecera, filas)

    # lista de ID_Pago tal y como quedaron en Pago_Pedido (columna correspondiente)
    col_id = letra_pp["ID_Pago"]
    ids = [ws_pp[f"{col_id}{r}"].value for r in range(2, ws_pp.max_row + 1)]
    escribir_resumen(wb, letra_pp, ids)

    fecha_str = datetime.now().strftime("%Y%m%d")
    destino = nombre_libre(ruta.parent, fecha_str)
    wb.save(destino)
    return destino, len(filas), len(set(ids))


# ----------------------------------------------------------------------
# Programa principal (GUI)
# ----------------------------------------------------------------------
def main():
    import tkinter as tk
    from tkinter import filedialog, messagebox

    raiz = tk.Tk()
    raiz.withdraw()

    archivo = filedialog.askopenfilename(
        title="Selecciona el extracto de pagos descargado de TikTok",
        filetypes=[
            ("Excel o CSV", "*.xlsx *.xls *.csv"),
            ("Todos los archivos", "*.*"),
        ],
    )
    if not archivo:
        return

    ruta = Path(archivo)
    try:
        destino, n_filas, n_pagos = procesar(ruta)
    except PermissionError:
        messagebox.showerror(
            "Error",
            "No se pudo guardar el archivo.\n\n"
            "Probablemente el Extracto esta abierto en Excel. Cierralo y reintenta.",
        )
        return
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo procesar el archivo:\n{e}")
        return

    messagebox.showinfo(
        "Listo",
        f"Archivo creado:\n{destino}\n\n"
        f"Hojas generadas:\n"
        f"  - Order details (copia)\n"
        f"  - Pago_Pedido (con columna CHECK)\n"
        f"  - Resumen_Pagos (consolidado por ID_Pago)\n\n"
        f"Lineas procesadas: {n_filas}\n"
        f"ID_Pago distintos: {n_pagos}\n\n"
        f"Revisa la columna CHECK: las filas 'REVISAR' no cuadran con el desglose.",
    )


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        import tkinter as tk
        from tkinter import messagebox
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("Error inesperado", str(e))
        sys.exit(1)